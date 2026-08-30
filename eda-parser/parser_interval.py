"""
EDA-Viertelstundenwerte-Parser ("Energiedaten"-Sheet)

Zweiter EDA-Export-Typ, komplett anders aufgebaut als der monatliche Energiedatenreport
(siehe parser.py) -- kein Zeilen-je-Zählpunkt-Format, sondern echte Viertelstundenwerte
(2.976 Zeitstempel für einen vollen Monat) in einem breiten Spaltenformat: jeder Zählpunkt
belegt mehrere Spaltenpaare (Wert + Datenqualität "MM"), ein Spaltenpaar je Kennzahl.

Format anhand einer echten Exportdatei verifiziert (Patrick, 03.09.2026,
RC108175_20260701T00_0020260731T23_45.xlsx):
- Spalte 1: Metadaten-Label je Zeile (MeteringPointId, Energydirection, Metering Interval, ...),
  danach ab der ersten erkennbaren Zeitstempel-Zeile die eigentlichen Viertelstunden-Zeitpunkte.
- Für jeden Zählpunkt mehrere Kennzahl-Spaltenpaare (Wert-Spalte + direkt danebenliegende
  "MM"-Qualitätsspalte), je nach Energierichtung:
    CONSUMPTION: Gesamtverbrauch, Verbrauch (Teilnahmefaktor), Anteil gem. Erzeugung,
                 Eigendeckung gem. Erzeugung, Eigendeckung aus erneuerbarer Energie
    GENERATION:  Gesamte gemeinschaftliche Erzeugung [kWh],
                 Erzeugung lt. Messung entsprechend dem Teilnahmefaktor und EC-ID [kWh],
                 Gesamt/Überschusserzeugung, Gemeinschaftsüberschuss [kWh] (Label KEIN Bindestrich
                 vor dem Schrägstrich -- anders als ursprünglich angenommen, siehe Fund 09.09.2026
                 anhand einer echten Exportdatei RC108175_20260801T00_0020260830T23_45.xlsx),
                 Restüberschuss bei EG und je ZP [kWh]
  Für die Mitglieder-Diagramme werden nur die jeweils ersten zwei Kennzahlen gebraucht (siehe
  TARGET_LABELS unten) -- die übrigen sind in der Datei vorhanden, aber (noch) nicht relevant.
- Am Ende der Zählpunkt-Spalten folgen "TOTAL"-Spalten (Community-Summe) -- werden übersprungen,
  nicht Teil der Zählpunkt-Zuordnung.
- Jeder Zählpunkt hat SEIN EIGENES "Data Period Start"/"-End" -- ein neu angeschlossener
  Zählpunkt kann also erst mitten im Exportzeitraum echte Werte haben (vorher: leere Zellen).
  Kennzahl-Spalten mit leerem Wert werden beim Import übersprungen, nicht als 0 gespeichert.

Aufruf:
  python parser_interval.py --file RC108175_..._QH.xlsx --community strompool-feldkirchen \
                             --user-id <uuid>
"""

import argparse
import json
import logging
import os
import re
from dataclasses import dataclass, field
from datetime import datetime

import openpyxl
import psycopg2
import psycopg2.extras

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

# Siehe parser.py für die Begründung (Batch-Prozess, kein direkter HTTP-Zugriff, gleiches Muster).
_DB_USER = os.environ.get("DB_USER", "eeg")
_DB_PASSWORD = os.environ.get("DB_PASSWORD", "")
DB_DSN = (
    f"host={os.environ.get('DB_HOST', 'localhost')} "
    f"port={os.environ.get('DB_PORT', '5432')} "
    f"dbname={os.environ.get('DB_NAME', 'eeg_platform')} "
    f"user={_DB_USER} "
    f"password={_DB_PASSWORD}"
)

SHEET_NAME = "Energiedaten"

# Welche Kennzahl-Spalte (per Label, robust gegen Reihenfolge-Änderungen) für die
# gespeicherten Werte je Energierichtung herangezogen wird -- siehe kwh_messung/kwh_gemeinschaft/
# kwh_erzeugung_gesamt in database/migrate_20260904.sql bzw. migrate_20260907.sql.
# "kwh_erzeugung_gesamt" nur bei GENERATION vorhanden (optional, siehe _find_target_col()-Aufruf
# unten) -- die eigene GESAMTE Erzeugung des Zählpunkts, im Gegensatz zu kwh_messung (dort
# gemeinschaftsweite Summe über alle Einspeiser) und kwh_gemeinschaft (nur der über den
# Teilnahmefaktor zugeteilte Anteil), siehe Kommentar in migrate_20260907.sql.
TARGET_LABELS = {
    "CONSUMPTION": {
        "kwh_messung": "gesamtverbrauch lt. messung",
        "kwh_gemeinschaft": "eigendeckung gemeinschaftliche erzeugung",
    },
    "GENERATION": {
        "kwh_messung": "gesamte gemeinschaftliche erzeugung",
        "kwh_gemeinschaft": "erzeugung lt. messung entsprechend dem teilnahmefaktor",
        "kwh_erzeugung_gesamt": "gesamt/überschusserzeugung",
    },
}

TIMESTAMP_RE = re.compile(r"^\d{2}\.\d{2}\.\d{4} \d{2}:\d{2}$")


@dataclass
class IntervalRow:
    time: datetime
    kwh_messung: float | None
    kwh_gemeinschaft: float | None
    kwh_erzeugung_gesamt: float | None
    quality: str | None


@dataclass
class MeteringPointInterval:
    zaehlpunkt_nr: str
    energy_direction: str
    rows: list[IntervalRow] = field(default_factory=list)


@dataclass
class LoadResult:
    metering_points: list[MeteringPointInterval]
    period_from: datetime
    period_to: datetime
    warnings: list[str] = field(default_factory=list)


def _worst_quality(a: str | None, b: str | None) -> str | None:
    for q in ("L3", "L2", "L1"):
        if a == q or b == q:
            return q
    return a or b


class IntervalXlsxDataSource:
    """Liest das "Energiedaten"-Sheet (Viertelstundenwerte) des EDA-Portals."""

    def load(self, filepath: str) -> LoadResult:
        log.info("Lese XLSX (Viertelstundenwerte): %s", filepath)
        # BEWUSST NICHT read_only=True: dieses Sheet wird spaltenweise wahlfrei gelesen (jede
        # Kennzahl-Spalte einzeln über den vollen Zeitstempel-Bereich), read_only-Worksheets sind
        # aber nur für sequenzielles Durchlaufen (iter_rows()) performant optimiert -- wahlfreier
        # Zell-Zugriff (ws.cell(row=, column=)) darauf war um Größenordnungen langsamer (>2 Min.
        # ohne Ergebnis bei einer 31-Tage-Datei, siehe Testlauf 03.09.2026).
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet_name = self._find_sheet(wb, SHEET_NAME)
        if not sheet_name:
            raise ValueError(
                f"XLSX hat nicht das erwartete Sheet '{SHEET_NAME}' (EDA-Viertelstundenwerte-"
                f"Export). Tatsächlich vorhandene Sheets: {wb.sheetnames}"
            )
        ws = wb[sheet_name]

        # Zeile-1-Labels (Spalte 1) einlesen, um die Metadaten-Zeilen dynamisch zu finden statt
        # feste Zeilennummern anzunehmen (robuster gegenüber künftigen Format-Anpassungen,
        # gleiches Prinzip wie _find_header_row() in parser.py).
        max_scan_rows = 40
        labels: dict[str, int] = {}
        data_start_row = None
        for r in range(1, max_scan_rows + 1):
            v = ws.cell(row=r, column=1).value
            if v is None:
                continue
            s = str(v).strip()
            if TIMESTAMP_RE.match(s):
                data_start_row = r
                break
            labels[s.lower()] = r
        if data_start_row is None:
            raise ValueError(
                "Konnte den Beginn der Viertelstunden-Zeitstempel in Spalte 1 nicht finden "
                f"(erste {max_scan_rows} Zeilen durchsucht)."
            )
        row_mpid = labels.get("meteringpointid")
        row_direction = labels.get("energydirection")
        row_metric_label = labels.get("metercode")  # trägt hier die Kennzahl-Beschreibung, kein echter Code
        row_report_start = labels.get("report filter start")
        row_report_end = labels.get("report filter end")
        if not all([row_mpid, row_direction, row_metric_label, row_report_start, row_report_end]):
            raise ValueError(
                "Erwartete Metadaten-Zeilen (MeteringPointId/Energydirection/MeterCode/"
                "Report Filter Start/-End) nicht vollständig in Spalte 1 gefunden."
            )

        max_col = ws.max_column
        # Reale letzte befüllte Spalte suchen -- die Datei ist oft mit leeren Spalten bis 1000
        # aufgefüllt, ein Scan bis dorthin wäre unnötig langsam.
        while max_col > 1 and ws.cell(row=row_mpid, column=max_col).value is None:
            max_col -= 1

        period_from = period_to = None
        # (zaehlpunkt_nr, energy_direction) -> {label_lower: (value_col, quality_col)}
        blocks: dict[tuple[str, str], dict[str, tuple[int, int]]] = {}
        order: list[tuple[str, str]] = []
        # log.warning() landet nur auf stderr -- bei einem ERFOLGREICHEN Lauf (Exit-Code 0,
        # gültiges JSON auf stdout) liest EdaParserRunner.php stderr zwar ein, verwirft es aber
        # ungeloggt (siehe dortiger Klassendoc: stderr wird nur bei einem FEHLGESCHLAGENEN Lauf
        # in die Diagnose übernommen). Jede hier auftretende Warnung deshalb zusätzlich in diese
        # Liste, die load() zurückgibt und import_to_db() in ihre eigene, tatsächlich sowohl im
        # Audit-Log als auch in der Import-Historie sichtbare "warnings"-Liste übernimmt (Fund
        # 09.09.2026: genau deshalb war eine fehlende Kennzahl-Spalte bisher nirgends sichtbar,
        # obwohl der Import selbst erfolgreich lief).
        warnings: list[str] = []

        c = 2
        while c <= max_col:
            mpid = ws.cell(row=row_mpid, column=c).value
            if mpid in (None, "MM"):
                c += 1
                continue
            mpid = str(mpid).strip()
            if mpid == "TOTAL":
                c += 2
                continue
            direction = ws.cell(row=row_direction, column=c).value
            label = ws.cell(row=row_metric_label, column=c).value
            if direction not in TARGET_LABELS or label is None:
                c += 2
                continue
            label_key = str(label).strip().lower()

            if period_from is None:
                pf = ws.cell(row=row_report_start, column=c).value
                pt = ws.cell(row=row_report_end, column=c).value
                if pf and pt:
                    period_from = datetime.strptime(str(pf).strip(), "%d.%m.%Y %H:%M:%S")
                    period_to = datetime.strptime(str(pt).strip(), "%d.%m.%Y %H:%M:%S")

            key = (mpid, direction)
            if key not in blocks:
                blocks[key] = {}
                order.append(key)
            blocks[key][label_key] = (c, c + 1)  # (Wert-Spalte, MM-Qualitätsspalte)
            c += 2

        if period_from is None or period_to is None:
            raise ValueError('„Report Filter Start"/"-End" im Datei-Kopf nicht gefunden.')

        # WICHTIG: ws.max_row NICHT in einer Schleifenbedingung wiederholt abfragen -- das ist
        # bei openpyxl kein gecachtes Attribut, sondern wird bei jedem Zugriff neu ermittelt.
        # Ein Testlauf mit `while r <= ws.max_row` brauchte dadurch >90s für einen einzigen
        # Zählpunkt (Test 03.09.2026); einmal vorab in eine lokale Variable gecacht, dauert der
        # komplette Import weniger als eine Sekunde.
        max_row = ws.max_row

        result = []
        for (mpid, direction), metric_cols in blocks.items():
            targets = TARGET_LABELS[direction]
            col_messung = self._find_target_col(metric_cols, targets["kwh_messung"])
            col_gemeinschaft = self._find_target_col(metric_cols, targets["kwh_gemeinschaft"])
            # Nur bei GENERATION vorhanden -- .get() statt [...], damit CONSUMPTION (kein
            # solcher Eintrag in TARGET_LABELS) nicht mit einem KeyError abbricht.
            col_erzeugung_gesamt = None
            erzeugung_gesamt_target = targets.get("kwh_erzeugung_gesamt")
            if erzeugung_gesamt_target:
                col_erzeugung_gesamt = self._find_target_col(metric_cols, erzeugung_gesamt_target)
                if col_erzeugung_gesamt is None:
                    # Anders als bei kwh_messung kein harter Abbruch (die Spalte ist optional,
                    # ältere Aufrufer/Zeilen kommen ohne sie aus) -- aber eine sichtbare Warnung,
                    # falls die tatsächliche EDA-Spaltenbeschriftung von der hier angenommenen
                    # ("gesamt-/überschusserzeugung", siehe TARGET_LABELS) abweicht, statt still
                    # jede Zeile ohne Gesamterzeugung zu importieren.
                    msg = (
                        f"Zählpunkt {mpid} ({direction}): Kennzahl-Spalte für kwh_erzeugung_gesamt "
                        f"nicht gefunden (gesucht: '{erzeugung_gesamt_target}', vorhanden: "
                        f"{sorted(metric_cols.keys())}) -- wird ohne eigene Gesamterzeugung importiert."
                    )
                    log.warning(msg)
                    warnings.append(msg)
            if col_messung is None:
                msg = f"Zählpunkt {mpid} ({direction}): Kennzahl-Spalte für kwh_messung nicht gefunden, übersprungen."
                log.warning(msg)
                warnings.append(msg)
                continue

            mp = MeteringPointInterval(zaehlpunkt_nr=mpid, energy_direction=direction)
            val_col, qual_col = col_messung
            gem_val_col = col_gemeinschaft[0] if col_gemeinschaft else None
            erzeugung_gesamt_val_col = col_erzeugung_gesamt[0] if col_erzeugung_gesamt else None

            r = data_start_row
            while r <= max_row:
                ts_raw = ws.cell(row=r, column=1).value
                if ts_raw is None:
                    break
                ts_str = str(ts_raw).strip()
                if not TIMESTAMP_RE.match(ts_str):
                    break
                wert = ws.cell(row=r, column=val_col).value
                if wert is not None:
                    qualitaet = ws.cell(row=r, column=qual_col).value
                    gemeinschaft = ws.cell(row=r, column=gem_val_col).value if gem_val_col else None
                    erzeugung_gesamt = ws.cell(row=r, column=erzeugung_gesamt_val_col).value if erzeugung_gesamt_val_col else None
                    mp.rows.append(IntervalRow(
                        time=datetime.strptime(ts_str, "%d.%m.%Y %H:%M"),
                        kwh_messung=float(wert),
                        kwh_gemeinschaft=float(gemeinschaft) if gemeinschaft is not None else None,
                        kwh_erzeugung_gesamt=float(erzeugung_gesamt) if erzeugung_gesamt is not None else None,
                        quality=str(qualitaet).strip() if qualitaet else None,
                    ))
                r += 1

            if mp.rows:
                result.append(mp)
            else:
                msg = f"Zählpunkt {mpid} ({direction}): keine Werte im Exportzeitraum, übersprungen."
                log.warning(msg)
                warnings.append(msg)

        wb.close()
        return LoadResult(metering_points=result, period_from=period_from, period_to=period_to, warnings=warnings)

    @staticmethod
    def _find_target_col(metric_cols: dict[str, tuple[int, int]], target_substr: str) -> tuple[int, int] | None:
        for label_key, cols in metric_cols.items():
            if target_substr in label_key:
                return cols
        return None

    @staticmethod
    def _find_sheet(wb: openpyxl.Workbook, expected_name: str) -> str | None:
        target = expected_name.strip().lower()
        for name in wb.sheetnames:
            if name.strip().lower() == target:
                return name
        return None


def import_to_db(
    conn,
    community_id: str,
    data: list[MeteringPointInterval],
    filename: str,
    user_id: str | None,
    period_from: datetime,
    period_to: datetime,
    load_warnings: list[str] | None = None,
) -> dict:
    """Schreibt die Viertelstundenwerte. Anders als beim monatlichen Import (parser.py) sind sich
    ÜBERLAPPENDE/wiederholte Zeiträume hier der Normalfall (Patrick lädt bewusst alle paar Tage
    einen neuen, teils überlappenden Ausschnitt hoch, siehe Kommentar in
    database/migrate_20260904.sql) -- deshalb kein Duplikat-Fehler wie in parser.py, sondern pro
    Zählpunkt ein einfaches "vorhandene Werte in genau diesem Zeitraum löschen, neu einfügen".
    Unbekannte Zählpunkte werden wie beim Monatsimport automatisch angelegt (inaktiv, ohne
    Mitglied-Zuordnung), damit ihre Daten nicht verloren gehen.
    load_warnings: Warnungen aus IntervalXlsxDataSource.load() (z.B. nicht gefundene
    Kennzahl-Spalten) -- fließen hier mit ein, damit sie über dieselbe "warnings"-Liste sowohl
    in eda_interval_imports als auch im Audit-Log/der UI landen, statt nur auf stderr zu verpuffen."""
    warnings: list[str] = list(load_warnings or [])
    neu_angelegt: list[dict] = []
    total_records = 0

    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("SELECT id, zaehlpunkt_nr FROM metering_points WHERE community_id = %s", (community_id,))
        registered = {row["zaehlpunkt_nr"]: str(row["id"]) for row in cur.fetchall()}

    with conn.cursor() as cur:
        for mp in data:
            if mp.zaehlpunkt_nr in registered:
                continue
            type_hint = "producer" if mp.energy_direction == "GENERATION" else "consumer"
            cur.execute(
                """
                INSERT INTO metering_points
                    (community_id, member_id, zaehlpunkt_nr, meter_code, type, active, registered_at)
                VALUES (%s, NULL, %s, '', %s, false, CURRENT_DATE)
                RETURNING id
                """,
                (community_id, mp.zaehlpunkt_nr, type_hint)
            )
            new_id = str(cur.fetchone()[0])
            registered[mp.zaehlpunkt_nr] = new_id
            neu_angelegt.append({"zaehlpunkt_nr": mp.zaehlpunkt_nr, "type_guess": type_hint, "metering_point_id": new_id})
            log.warning("Zählpunkt %s automatisch angelegt (Typ-Vermutung: %s, unzugeordnet)", mp.zaehlpunkt_nr, type_hint)

    if neu_angelegt:
        warnings.append(f"{len(neu_angelegt)} neu angelegte, noch nicht zugeordnete Zählpunkte.")

    with conn.cursor() as cur:
        for mp in data:
            mp_id = registered.get(mp.zaehlpunkt_nr)
            if not mp_id or not mp.rows:
                continue
            ts_min = min(row.time for row in mp.rows)
            ts_max = max(row.time for row in mp.rows)

            cur.execute(
                "DELETE FROM eda_interval_data WHERE community_id = %s AND metering_point_id = %s AND time >= %s AND time <= %s",
                (community_id, mp_id, ts_min, ts_max)
            )

            rows = [
                (row.time, community_id, mp_id, mp.energy_direction, row.kwh_messung, row.kwh_gemeinschaft,
                 row.kwh_erzeugung_gesamt, row.quality)
                for row in mp.rows
            ]
            psycopg2.extras.execute_values(
                cur,
                """
                INSERT INTO eda_interval_data
                    (time, community_id, metering_point_id, energy_direction, kwh_messung, kwh_gemeinschaft,
                     kwh_erzeugung_gesamt, quality)
                VALUES %s
                """,
                rows,
            )
            total_records += len(rows)
            log.info("Zählpunkt %s: %d Viertelstundenwerte importiert (%s – %s)", mp.zaehlpunkt_nr, len(rows), ts_min, ts_max)

        cur.execute(
            """
            INSERT INTO eda_interval_imports (community_id, imported_by, filename, period_from, period_to, records_imported, warnings)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id
            """,
            (community_id, user_id, filename, period_from, period_to, total_records, json.dumps(warnings))
        )
        import_id = cur.fetchone()[0]

    conn.commit()
    log.info("Import abgeschlossen: %d Datensätze, %d Warnungen, %d neu angelegt (Import-ID: %s)",
              total_records, len(warnings), len(neu_angelegt), import_id)

    return {
        "import_id": str(import_id),
        "records": total_records,
        "warnings": warnings,
        "neu_angelegt": neu_angelegt,
        "period_from": str(period_from),
        "period_to": str(period_to),
    }


def main():
    parser = argparse.ArgumentParser(description="EDA-Viertelstundenwerte-Importer (Energiedaten-Sheet)")
    parser.add_argument("--file", required=True, help="Pfad zur XLSX-Datei")
    parser.add_argument("--community", required=True, help="Community-Slug")
    parser.add_argument("--user-id", help="UUID des importierenden Users")
    args = parser.parse_args()

    conn = psycopg2.connect(DB_DSN)
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM communities WHERE slug = %s", (args.community,))
            row = cur.fetchone()
            if not row:
                raise SystemExit(f"Community '{args.community}' nicht gefunden")
            community_id = str(row[0])

        source = IntervalXlsxDataSource()
        loaded = source.load(args.file)
        result = import_to_db(
            conn, community_id, loaded.metering_points, os.path.basename(args.file), args.user_id,
            loaded.period_from, loaded.period_to, load_warnings=loaded.warnings,
        )
        print(json.dumps(result, indent=2, ensure_ascii=False))
    finally:
        conn.close()


if __name__ == "__main__":
    main()
